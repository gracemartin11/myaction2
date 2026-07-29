#!/usr/bin/env python3
"""
Bluesky auto-poster.

Coordination data (Credentials / Settings / PostPlan / LinkPlan / Report)
lives in a SINGLE .xlsx WORKBOOK stored on Mega (via rclone) — not Google
Sheets. Because Mega has no per-cell API, every write means "download the
whole file, change it, upload the whole file back". To make that safe across
several parallel account jobs, all reads-that-lead-to-writes are wrapped in
one real distributed lock:

    1. Try to atomically claim a lock marker on Mega with `rclone moveto`
       (the exact same atomic-rename trick already used to claim media
       files below — a moveto either fully succeeds or fully fails, so two
       runners can never both "win" the same rename).
    2. Once the lock is held: download the freshest copy of the workbook,
       make the change, upload it, then move the lock marker back
       (release). This whole cycle is kept as short as possible — no
       Bluesky API calls, HTTP fetches, or file downloads ever happen while
       the lock is held.
    3. If a runner crashes while holding the lock, the marker is left
       "held" forever, so lock acquisition also does TTL-based staleness
       detection: if a held marker is older than WORKBOOK_LOCK_TTL_SECONDS,
       any other runner may attempt to steal it (again via an atomic
       moveto — only one thief can ever succeed).

This is a lot more moving parts than a real database transaction, and it is
NOT as robust as using an actual application with row-level locking (a
real DB, or Google Sheets' per-cell API) would be — concurrent writes are
still serialized through a single file, so throughput is limited and a
sufficiently badly-timed crash mid-upload could still corrupt the workbook.
But it removes the Google Sheets dependency entirely and keeps everything,
including all coordination state, on Mega.

Media files (images/videos) still live on Mega, accessed via rclone,
unchanged from before.
"""
import io
import json
import os
import random
import re
import socket
import subprocess
import sys
import time
import uuid
from datetime import datetime
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from atproto import Client, models
from atproto_client.utils import TextBuilder
import openpyxl

RUN_TAG      = os.getenv("GITHUB_RUN_ID") or f"{socket.gethostname()}-{uuid.uuid4().hex[:8]}"
CLAIM_PREFIX = "CLAIMED_"

CURRENT_REPO     = os.getenv("GITHUB_REPOSITORY") or f"local-{socket.gethostname()}"
LOCK_TTL_MINUTES = 45   # business-level "which repo owns this account row" heartbeat TTL

# ═══════════════════════════════════════════════════════════════════════════
#  ENV / VALUE PARSING HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def get_env(name, required=True, default=""):
    v = os.getenv(name)
    if v is None or not v.strip():
        if required:
            raise RuntimeError(f"Missing required env var: {name}")
        return default
    return v.strip()

def _parse_bool(raw, default=False):
    if raw is None or not str(raw).strip():
        return default
    return str(raw).strip().lower() in ("1", "true", "yes", "on")

def _parse_pct(raw, default):
    if raw is None or not str(raw).strip():
        return default
    raw = str(raw).strip().rstrip("%")
    try:
        v = float(raw)
        return v / 100.0 if v > 1 else v
    except ValueError:
        return default

def _parse_plain_float(raw, default):
    if raw is None or not str(raw).strip():
        return default
    try:
        return float(str(raw).strip())
    except ValueError:
        return default

def _parse_int(raw, default):
    if raw is None or not str(raw).strip():
        return default
    try:
        return max(1, int(str(raw).strip()))
    except ValueError:
        return default

def get_bool_env(name, default=False):
    return _parse_bool(os.getenv(name), default)

def get_float_env(name, default):
    return _parse_pct(os.getenv(name), default)

def get_int_env(name, default):
    return _parse_int(os.getenv(name), default)


# ═══════════════════════════════════════════════════════════════════════════
#  STATIC WORKFLOW KNOBS
# ═══════════════════════════════════════════════════════════════════════════

ACCOUNT_ROW = get_int_env("ACCOUNT_ROW", 1)   # 1-based data row (header is row 1 in the sheet)

RCLONE_CONFIG_PATH = get_env("RCLONE_CONFIG_PATH", required=False) or "rclone.conf"
RCLONE_REMOTE_NAME = get_env("RCLONE_REMOTE_NAME", required=False) or "mega"

CREDS_TAB    = "Credentials"
SETTINGS_TAB = "Settings"
REPORT_TAB   = "Report"

SKIP_STATUS_MARKERS = ("banned", "suspended", "taken down", "auth failed")

REPORT_HEADER = ["Timestamp (UTC)", "Handle", "Followers", "Gained", "Top Post", "Engagement", "Status"]

CREDENTIALS_HEADER = [
    "BSKY_HANDLE", "BSKY_APP_PW", "HASHTAGS",
    "MEGA_UPLOAD_FOLDER", "MEGA_PROCESSED_FOLDER",
    "LINK_URL", "LINK_DISPLAY_TEXT",
    "LOCKED_BY", "LOCKED_AT",
    "ACCOUNT_STATUS", "ACCOUNT_STATUS_AT",
    "ASSIGNED_REPO", "ASSIGNED_STATUS", "ASSIGNED_AT",
]

DEFAULT_SETTINGS = [
    ("IMAGE_RATIO", "0.60"),
    ("VIDEO_RATIO", "0.40"),
    ("LINK_RATIO", "0.0"),
    ("HASHTAGS_ENABLED_IMAGE", "true"),
    ("HASHTAGS_ENABLED_VIDEO", "false"),
    ("HASHTAGS_ENABLED_LINK", "true"),
    ("LINK_ENABLED_IMAGE", "true"),
    ("LINK_ENABLED_VIDEO", "true"),
    ("LINK_PERCENTAGE", "1.0"),
    ("MAX_IMAGE_MB", "2.0"),
    ("CAPTION_ENABLED", "true"),
    ("AUTO_CAPTION_ENABLED_LINK", "true"),
    ("PREVIEW_FETCH_TIMEOUT", "15"),
    ("MAX_THUMB_MB", "1.0"),
    ("ENABLE_REPORT", "false"),
    ("REPORT_TIMES_PER_DAY", "1"),
    ("TOP_POSTS_COUNT", "1"),
    ("TOP_POSTS_WITHIN", "30"),
    ("POST_PLAN_SHEET_NAME", "PostPlan"),
    ("LINK_PLAN_SHEET_NAME", "LinkPlan"),
    ("LOOP_INTERVAL_SECONDS", "1800"),
    ("MAX_ACCOUNTS_PER_RUN", ""),
]

POSTED_STATUS_VALUE = "posted"
ASSIGN_STATUS_IN_USE = "In Use"

_URL_RE     = re.compile(r"https?://\S+")
_MENTION_RE = re.compile(r"@\S+")

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

CARDYB_EXTRACT_URL = "https://cardyb.bsky.app/v1/extract"
LINK_PREVIEW_MAX_RETRIES = 3
LINK_PREVIEW_RETRY_DELAY = 2

DEFAULT_LOOP_INTERVAL_SECONDS = 1800


# ═══════════════════════════════════════════════════════════════════════════
#  RCLONE HELPERS (used for both media files and the state workbook/lock)
# ═══════════════════════════════════════════════════════════════════════════

def _rclone_run(args, timeout=120):
    return subprocess.run(
        ["rclone", "--config", RCLONE_CONFIG_PATH] + args,
        capture_output=True, text=True, timeout=timeout,
    )

def rclone_list_files(remote_folder):
    result = _rclone_run(["lsf", remote_folder, "--files-only"])
    if result.returncode != 0:
        print(f"Warning: rclone lsf failed for '{remote_folder}': {result.stderr.strip()[-300:]}")
        return []
    return [line.strip() for line in result.stdout.splitlines() if line.strip()]

def rclone_lsjson(remote_folder):
    result = _rclone_run(["lsjson", remote_folder])
    if result.returncode != 0:
        return []
    try:
        return json.loads(result.stdout)
    except Exception:
        return []

def rclone_claim(remote_folder, name):
    claimed_name = f"{CLAIM_PREFIX}{RUN_TAG}__{name}"
    result = _rclone_run(["moveto", f"{remote_folder}/{name}", f"{remote_folder}/{claimed_name}"])
    return claimed_name if result.returncode == 0 else None

def rclone_download(remote_folder, filename, local_path, timeout=120):
    result = _rclone_run(["copyto", f"{remote_folder}/{filename}", local_path], timeout=timeout)
    return result.returncode == 0

def rclone_move(src, dst, timeout=120):
    result = _rclone_run(["moveto", src, dst], timeout=timeout)
    return result.returncode == 0

def rclone_copyto(src, dst, timeout=120):
    result = _rclone_run(["copyto", src, dst], timeout=timeout)
    return result.returncode == 0


# ═══════════════════════════════════════════════════════════════════════════
#  MEGA STATE-WORKBOOK LOCATION
# ═══════════════════════════════════════════════════════════════════════════

# Path to the single .xlsx workbook on Mega that holds Credentials / Settings /
# Report / PostPlan / LinkPlan, e.g. "AutoPoster/state.xlsx".
MEGA_STATE_PATH        = get_env("MEGA_STATE_PATH")
MEGA_STATE_REMOTE_NAME = get_env("MEGA_STATE_REMOTE_NAME", required=False) or RCLONE_REMOTE_NAME

WORKBOOK_LOCK_TTL_SECONDS             = get_int_env("WORKBOOK_LOCK_TTL_SECONDS", 300)
WORKBOOK_LOCK_ACQUIRE_TIMEOUT_SECONDS = get_int_env("WORKBOOK_LOCK_ACQUIRE_TIMEOUT_SECONDS", 240)
WORKBOOK_LOCK_POLL_SECONDS            = get_int_env("WORKBOOK_LOCK_POLL_SECONDS", 4)

STATE_LOCAL_PATH = "/tmp/_state_workbook.xlsx"


def _remote_join(remote_dir, name):
    """Join an rclone remote directory (e.g. 'mega:folder' or 'mega:') with a filename."""
    if remote_dir.endswith(":") or remote_dir.endswith("/"):
        return f"{remote_dir}{name}"
    return f"{remote_dir}/{name}"


def _split_remote_path(remote_name, path):
    path = path.strip("/")
    if "/" in path:
        d, f = path.rsplit("/", 1)
        return f"{remote_name}:{d}", f
    return f"{remote_name}:", path


STATE_REMOTE_DIR, STATE_FILENAME = _split_remote_path(MEGA_STATE_REMOTE_NAME, MEGA_STATE_PATH)
STATE_REMOTE_FULL = _remote_join(STATE_REMOTE_DIR, STATE_FILENAME)
LOCK_FILENAME     = STATE_FILENAME + ".lock"           # marker name when the lock is FREE
LOCK_HELD_PREFIX  = LOCK_FILENAME + ".held."           # marker prefix + RUN_TAG when HELD


# ═══════════════════════════════════════════════════════════════════════════
#  REAL DISTRIBUTED LOCK ON MEGA (atomic rename via `rclone moveto`)
# ═══════════════════════════════════════════════════════════════════════════

_RCLONE_TS_RE = re.compile(r"^(\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2})(\.\d+)?(Z|[+-]\d{2}:\d{2})?$")

def _parse_rclone_time(s):
    if not s:
        return None
    m = _RCLONE_TS_RE.match(s.strip())
    if not m:
        return None
    base, _frac, tz = m.groups()
    iso = base + (tz if tz and tz != "Z" else "+00:00")
    try:
        return datetime.fromisoformat(iso).timestamp()
    except Exception:
        return None


class MegaLock:
    """A real cross-process lock backed by an atomic Mega file rename.

    Steady state: exactly one file exists, either
      - <name>.lock              (FREE), or
      - <name>.lock.held.<tag>   (HELD by whichever runner's tag is in the name)

    Acquiring = `rclone moveto <free> <held-by-me>`. This is atomic: if two
    runners race, only one `moveto` can possibly succeed (the source
    disappears the instant the winner's rename lands), so there is no
    check-then-act gap for a normal handoff.

    If a runner dies while holding the lock, the `.held.<tag>` marker is
    orphaned forever. Any other runner may steal an orphaned marker once
    it's older than `ttl_seconds`, again via an atomic moveto — so at most
    one thief ever wins that race too.
    """

    def __init__(self, remote_dir, lock_filename, owner, ttl_seconds, acquire_timeout, poll_interval):
        self.remote_dir  = remote_dir.rstrip("/") if not remote_dir.endswith(":") else remote_dir
        self.free_path   = _remote_join(self.remote_dir, lock_filename)
        self.held_prefix = f"{lock_filename}.held."
        self.owner       = owner
        self.ttl         = ttl_seconds
        self.timeout     = acquire_timeout
        self.poll        = poll_interval
        self._held_path  = None

    def _held_path_for(self, tag):
        return _remote_join(self.remote_dir, f"{self.held_prefix}{tag}")

    def _find_held_entry(self):
        for item in rclone_lsjson(self.remote_dir):
            name = item.get("Name", "")
            if name.startswith(self.held_prefix):
                return item
        return None

    def _bootstrap_if_nothing_exists(self):
        """One-time, best-effort: if neither a free nor a held marker exists
        anywhere (i.e. this is the very first run ever), create the free
        marker so normal acquire/steal logic has something to work with.
        There's an unavoidable small race window here on the true first run
        only — recommend triggering one manual run before turning on the
        schedule to get past it safely."""
        entries = rclone_lsjson(self.remote_dir)
        names = {e.get("Name") for e in entries}
        lock_filename = self.free_path.rsplit("/", 1)[-1]
        has_free = lock_filename in names
        has_held = any(n.startswith(self.held_prefix) for n in names)
        if not has_free and not has_held:
            tmp = f"/tmp/_lockinit_{uuid.uuid4().hex}"
            open(tmp, "w").close()
            rclone_copyto(tmp, self.free_path)
            try:
                os.remove(tmp)
            except OSError:
                pass

    def acquire(self):
        deadline = time.time() + self.timeout
        attempt = 0
        while True:
            attempt += 1
            my_held = self._held_path_for(self.owner)

            if rclone_move(self.free_path, my_held):
                self._held_path = my_held
                return True

            entry = self._find_held_entry()
            if entry:
                ts = _parse_rclone_time(entry.get("ModTime"))
                stale = ts is None or (time.time() - ts) > self.ttl
                if stale:
                    stale_path = _remote_join(self.remote_dir, entry["Name"])
                    if rclone_move(stale_path, my_held):
                        print(f"[workbook-lock] reclaimed stale lock '{entry['Name']}' "
                              f"(idle > {self.ttl}s).")
                        self._held_path = my_held
                        return True
            else:
                self._bootstrap_if_nothing_exists()

            if time.time() >= deadline:
                raise TimeoutError(
                    f"Could not acquire the workbook lock within {self.timeout}s "
                    f"(another runner is holding it and it isn't stale yet)."
                )
            wait = self.poll + random.uniform(0, self.poll)
            print(f"[workbook-lock] attempt {attempt}: busy, retrying in {wait:.1f}s…")
            time.sleep(wait)

    def release(self):
        if not self._held_path:
            return
        if not rclone_move(self._held_path, self.free_path):
            print("Warning: could not cleanly release the workbook lock — "
                  "another runner may have force-reclaimed it as stale. "
                  "This is recoverable but worth keeping an eye on.")
        self._held_path = None

    def __enter__(self):
        self.acquire()
        return self

    def __exit__(self, exc_type, exc, tb):
        self.release()
        return False


# ═══════════════════════════════════════════════════════════════════════════
#  WORKBOOK-LEVEL HELPERS (openpyxl worksheet access, 1-based like the sheet was)
# ═══════════════════════════════════════════════════════════════════════════

def ws_header_map(ws):
    header = {}
    if ws.max_row >= 1:
        for i, cell in enumerate(ws[1], start=1):
            v = cell.value
            if v is not None and str(v).strip():
                header[str(v).strip().upper()] = i
    return header

def ws_cell(ws, row, col):
    if col is None:
        return ""
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""

def ws_set_cell(ws, row, col, value):
    ws.cell(row=row, column=col, value=value)


REQUIRED_TABS = {
    CREDS_TAB:    CREDENTIALS_HEADER,
    SETTINGS_TAB: ["KEY", "VALUE"],
    REPORT_TAB:   REPORT_HEADER,
}


class WorkbookStore:
    """The single .xlsx workbook on Mega, accessed either as:
      - snapshot(): an unlocked, best-effort-fresh read-only download. Fine
        for pure reads (settings, scanning PostPlan/LinkPlan) where a lost
        update isn't possible because nothing gets written back.
      - transact(fn): acquire the distributed lock, download the freshest
        copy, run fn(wb) -> (dirty: bool, result), upload only if dirty,
        release the lock. This is the ONLY path allowed to write, and it's
        kept as short as possible on purpose — no network calls other than
        the download/upload happen while the lock is held.
    """

    def __init__(self):
        self.lock = MegaLock(
            STATE_REMOTE_DIR, LOCK_FILENAME, owner=RUN_TAG,
            ttl_seconds=WORKBOOK_LOCK_TTL_SECONDS,
            acquire_timeout=WORKBOOK_LOCK_ACQUIRE_TIMEOUT_SECONDS,
            poll_interval=WORKBOOK_LOCK_POLL_SECONDS,
        )
        self.wb = None
        self._dirty = False

    def _download(self):
        ok = rclone_download(STATE_REMOTE_DIR, STATE_FILENAME, STATE_LOCAL_PATH)
        created_new = not (ok and os.path.exists(STATE_LOCAL_PATH))
        if created_new:
            # Build fresh in-memory rather than saving/reloading an empty
            # workbook — openpyxl refuses to save a workbook with zero
            # sheets, and we haven't added the required tabs yet.
            self.wb = openpyxl.Workbook()
            self.wb.remove(self.wb.active)
            print(f"Note: '{STATE_REMOTE_FULL}' doesn't exist yet on Mega — "
                  f"starting a brand-new workbook.")
        else:
            self.wb = openpyxl.load_workbook(STATE_LOCAL_PATH)
        self._dirty = created_new
        self._ensure_required_tabs()

    def _ensure_required_tabs(self):
        for tab, header in REQUIRED_TABS.items():
            if tab not in self.wb.sheetnames:
                ws = self.wb.create_sheet(tab)
                ws.append(header)
                self._dirty = True
            else:
                ws = self.wb[tab]
                if ws.max_row < 1 or all((c.value in (None, "")) for c in ws[1]):
                    for i, h in enumerate(header, start=1):
                        ws.cell(row=1, column=i, value=h)
                    self._dirty = True

    def _upload(self):
        self.wb.save(STATE_LOCAL_PATH)
        if not rclone_copyto(STATE_LOCAL_PATH, STATE_REMOTE_FULL):
            raise RuntimeError(
                "Failed to upload the workbook back to Mega — the lock will "
                "still be released, but this write did not make it, so a "
                "retry (of whatever operation this was) is a good idea."
            )

    def snapshot(self):
        self._download()
        return self.wb

    def transact(self, fn):
        with self.lock:
            self._download()
            result = fn(self.wb)
            if isinstance(result, tuple) and len(result) == 2 and isinstance(result[0], bool):
                dirty, ret = result
            else:
                dirty, ret = True, result
            if dirty or self._dirty:
                self._upload()
            return ret


STORE = WorkbookStore()


def bootstrap_state():
    """Force one transaction so the workbook file + required tabs definitely
    exist on Mega, even on a completely fresh setup."""
    STORE.transact(lambda wb: (True, None))


# ═══════════════════════════════════════════════════════════════════════════
#  ACCOUNT CONFIG + LIVE SETTINGS
# ═══════════════════════════════════════════════════════════════════════════

_account_config         = None
_global_settings_cache  = None


def load_global_settings(force_refresh=False):
    global _global_settings_cache
    if _global_settings_cache is not None and not force_refresh:
        return _global_settings_cache

    settings = {}
    try:
        wb = STORE.snapshot()
        ws = wb[SETTINGS_TAB]
        for r in range(2, ws.max_row + 1):
            key = ws_cell(ws, r, 1)
            val = ws_cell(ws, r, 2)
            if key:
                settings[key.upper()] = val
    except Exception as exc:
        print(f"Note: '{SETTINGS_TAB}' unreadable — using built-in defaults ({exc}).")

    _global_settings_cache = settings
    return _global_settings_cache


def ensure_settings_defaults():
    def _mut(wb):
        ws = wb[SETTINGS_TAB]
        existing = set()
        for r in range(2, ws.max_row + 1):
            k = ws_cell(ws, r, 1)
            if k:
                existing.add(k.upper())
        dirty = False
        for k, v in DEFAULT_SETTINGS:
            if k not in existing:
                ws.append([k, v])
                dirty = True
        return dirty, None
    STORE.transact(_mut)


def load_account_config(force_refresh=False):
    global _account_config

    if _account_config is not None and not force_refresh:
        return _account_config

    wb = STORE.snapshot()
    ws = wb[CREDS_TAB]
    header = ws_header_map(ws)

    excel_row = ACCOUNT_ROW + 1
    if excel_row > ws.max_row:
        raise RuntimeError(
            f"ACCOUNT_ROW={ACCOUNT_ROW} but '{CREDS_TAB}' only has "
            f"{max(0, ws.max_row - 1)} data row(s)."
        )

    def col(*names):
        for n in names:
            c = header.get(n.upper())
            if c is not None:
                return ws_cell(ws, excel_row, c)
        return ""

    shared = load_global_settings(force_refresh)
    def setting(key):
        return col(key) or shared.get(key, "")

    raw_link     = col("LINK_URL") or "https://foodiesposts.com"
    link_url     = raw_link if raw_link.startswith("http") else f"https://{raw_link}"
    link_display = col("LINK_DISPLAY_TEXT") or link_url.replace("https://", "").replace("http://", "")

    img_ratio_raw  = _parse_pct(setting("IMAGE_RATIO"), 0.60)
    vid_ratio_raw  = _parse_pct(setting("VIDEO_RATIO"), 0.40)
    link_ratio_raw = _parse_pct(setting("LINK_RATIO"), 0.0)
    ratio_sum      = img_ratio_raw + vid_ratio_raw + link_ratio_raw
    if ratio_sum > 0:
        image_ratio = img_ratio_raw / ratio_sum
        video_ratio = vid_ratio_raw / ratio_sum
        link_ratio  = link_ratio_raw / ratio_sum
    else:
        image_ratio, video_ratio, link_ratio = 0.60, 0.40, 0.0

    cfg = {
        "handle":                col("BSKY_HANDLE"),
        "app_pw":                col("BSKY_APP_PW"),
        "link_url":              link_url,
        "link_display_text":     link_display,
        "hashtags_raw":          col("HASHTAGS"),
        "mega_upload_folder":    col("MEGA_UPLOAD_FOLDER"),
        "mega_processed_folder": col("MEGA_PROCESSED_FOLDER"),
        "row_num":               ACCOUNT_ROW,

        "image_ratio":              image_ratio,
        "video_ratio":              video_ratio,
        "link_ratio":               link_ratio,
        "hashtags_enabled_image":   _parse_bool(setting("HASHTAGS_ENABLED_IMAGE"), True),
        "hashtags_enabled_video":   _parse_bool(setting("HASHTAGS_ENABLED_VIDEO"), False),
        "hashtags_enabled_link":    _parse_bool(setting("HASHTAGS_ENABLED_LINK"), True),
        "link_enabled_image":       _parse_bool(setting("LINK_ENABLED_IMAGE"), True),
        "link_enabled_video":       _parse_bool(setting("LINK_ENABLED_VIDEO"), True),
        "link_percentage":          _parse_pct(setting("LINK_PERCENTAGE"), 1.0),
        "max_image_bytes":          int(_parse_plain_float(setting("MAX_IMAGE_MB"), 2.0) * 1024 * 1024),
        "caption_enabled":          _parse_bool(setting("CAPTION_ENABLED"), True),
        "auto_caption_enabled_link": _parse_bool(setting("AUTO_CAPTION_ENABLED_LINK"), True),
        "preview_timeout":          _parse_int(setting("PREVIEW_FETCH_TIMEOUT"), 15),
        "max_thumb_bytes":          int(_parse_plain_float(setting("MAX_THUMB_MB"), 1.0) * 1024 * 1024),
        "enable_report":            _parse_bool(setting("ENABLE_REPORT"), False),
        "report_times_per_day":     _parse_int(setting("REPORT_TIMES_PER_DAY"), 1),
        "top_posts_count":          _parse_int(setting("TOP_POSTS_COUNT"), 1),
        "top_posts_within":         _parse_int(setting("TOP_POSTS_WITHIN"), 30),
        "post_plan_sheet_name":     setting("POST_PLAN_SHEET_NAME") or "PostPlan",
        "link_plan_sheet_name":     setting("LINK_PLAN_SHEET_NAME") or "LinkPlan",
        "loop_interval_seconds":    _parse_int(setting("LOOP_INTERVAL_SECONDS"),
                                                DEFAULT_LOOP_INTERVAL_SECONDS),

        "locked_by": col("LOCKED_BY"),
        "locked_at": col("LOCKED_AT"),
        "account_status": col("ACCOUNT_STATUS"),
        "has_lock_columns": ("LOCKED_BY" in header and "LOCKED_AT" in header),
    }

    if not cfg["handle"]:
        raise RuntimeError(
            f"BSKY_HANDLE is empty for account row {ACCOUNT_ROW} in '{CREDS_TAB}'."
        )

    _account_config = cfg
    return cfg

def _cfg():
    return load_account_config()

def refresh_account_config():
    return load_account_config(force_refresh=True)


# ═══════════════════════════════════════════════════════════════════════════
#  AUTO ACCOUNT-ROW ASSIGNMENT (one short transaction: scan + claim)
# ═══════════════════════════════════════════════════════════════════════════

def resolve_account_row():
    explicit = get_env("ACCOUNT_ROW", required=False)
    if explicit:
        row = _parse_int(explicit, 1)
        print(f"ACCOUNT_ROW={row} was explicitly set — using it as a manual override "
              f"(auto-assignment skipped).")
        return row

    def _mut(wb):
        ws = wb[CREDS_TAB]
        header = ws_header_map(ws)

        def hidx(*names):
            for n in names:
                if n.upper() in header:
                    return header[n.upper()]
            return None

        handle_col = hidx("BSKY_HANDLE")
        repo_col   = hidx("ASSIGNED_REPO")
        status_col = hidx("ASSIGNED_STATUS")
        at_col     = hidx("ASSIGNED_AT")

        if handle_col is None or repo_col is None or status_col is None:
            raise RuntimeError(
                f"Auto row-assignment needs 'BSKY_HANDLE', 'ASSIGNED_REPO' and "
                f"'ASSIGNED_STATUS' columns in '{CREDS_TAB}'. Add any missing ones "
                f"to the header row, or set ACCOUNT_ROW manually for this run."
            )

        max_row = ws.max_row

        for excel_row in range(2, max_row + 1):
            if ws_cell(ws, excel_row, repo_col) == CURRENT_REPO:
                handle = ws_cell(ws, excel_row, handle_col)
                data_idx = excel_row - 1
                print(f"Repo '{CURRENT_REPO}' already owns Credentials row {data_idx} "
                      f"({handle or 'no handle'}) — reusing it.")
                return False, data_idx

        for excel_row in range(2, max_row + 1):
            handle_val = ws_cell(ws, excel_row, handle_col)
            status_val = ws_cell(ws, excel_row, status_col)
            if not handle_val:
                continue
            if status_val.lower() == ASSIGN_STATUS_IN_USE.lower():
                continue

            now = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
            ws_set_cell(ws, excel_row, repo_col, CURRENT_REPO)
            ws_set_cell(ws, excel_row, status_col, ASSIGN_STATUS_IN_USE)
            if at_col is not None:
                ws_set_cell(ws, excel_row, at_col, now)
            data_idx = excel_row - 1
            print(f"Claimed Credentials row {data_idx} ({handle_val}) for repo '{CURRENT_REPO}'.")
            return True, data_idx

        raise RuntimeError(
            f"No available account rows left in '{CREDS_TAB}' — every configured "
            f"row is already marked '{ASSIGN_STATUS_IN_USE}'. Add a new account "
            f"row, or clear ASSIGNED_REPO/ASSIGNED_STATUS on one you want to free up."
        )

    return STORE.transact(_mut)


# ═══════════════════════════════════════════════════════════════════════════
#  CROSS-REPO SOFT LOCK (business-level "who owns this account row right now")
# ═══════════════════════════════════════════════════════════════════════════

class AccountLockedElsewhereError(Exception):
    """Non-fatal — another repo currently owns this account row."""


def try_acquire_account_lock():
    global _account_config

    def _mut(wb):
        ws = wb[CREDS_TAB]
        header = ws_header_map(ws)
        by_col = header.get("LOCKED_BY")
        at_col = header.get("LOCKED_AT")
        excel_row = ACCOUNT_ROW + 1

        if by_col is None or at_col is None:
            return False, True   # no lock columns configured -> treat as always acquired

        locked_by     = ws_cell(ws, excel_row, by_col)
        locked_at_raw = ws_cell(ws, excel_row, at_col)

        stale = True
        if locked_at_raw:
            try:
                locked_at = time.mktime(time.strptime(locked_at_raw, "%Y-%m-%dT%H:%M:%SZ"))
                stale = (time.time() - locked_at) > LOCK_TTL_MINUTES * 60
            except ValueError:
                stale = True

        if locked_by and locked_by != CURRENT_REPO and not stale:
            print(f"Row {ACCOUNT_ROW} is currently locked by '{locked_by}' "
                  f"(last heartbeat {locked_at_raw} UTC, TTL {LOCK_TTL_MINUTES}m). Skipping this run.")
            return False, False

        now = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
        ws_set_cell(ws, excel_row, by_col, CURRENT_REPO)
        ws_set_cell(ws, excel_row, at_col, now)
        if _account_config:
            _account_config["locked_by"] = CURRENT_REPO
            _account_config["locked_at"] = now
        return True, True

    refresh_account_config()
    return STORE.transact(_mut)


def _write_account_status(status):
    global _account_config

    def _mut(wb):
        ws = wb[CREDS_TAB]
        header = ws_header_map(ws)
        status_col = header.get("ACCOUNT_STATUS")
        if status_col is None:
            print("Note: no ACCOUNT_STATUS column in Credentials — add one to track per-row status.")
            return False, None
        at_col = header.get("ACCOUNT_STATUS_AT")
        excel_row = ACCOUNT_ROW + 1
        ws_set_cell(ws, excel_row, status_col, status)
        if at_col is not None:
            ws_set_cell(ws, excel_row, at_col, _now_str())
        return True, None

    try:
        STORE.transact(_mut)
        if _account_config:
            _account_config["account_status"] = status
        print(f"Credentials ACCOUNT_STATUS set to '{status}' for row {ACCOUNT_ROW}.")
    except Exception as exc:
        print(f"Warning: could not update ACCOUNT_STATUS: {exc}")


# ═══════════════════════════════════════════════════════════════════════════
#  TEXT HELPERS
# ═══════════════════════════════════════════════════════════════════════════

def _posting_handle():
    h = _cfg()["handle"]
    return h if h.startswith("@") else f"@{h}"

def replace_mentions(text):
    return _MENTION_RE.sub(_posting_handle(), text) if text else text

def replace_urls(text):
    return _URL_RE.sub(_cfg()["link_url"], text) if text else text


# ═══════════════════════════════════════════════════════════════════════════
#  CONFIG SUMMARY
# ═══════════════════════════════════════════════════════════════════════════

def print_config_summary():
    cfg = _cfg()
    print("── Run config (live from Settings + Credentials in the Mega workbook) ──")
    print(f"  Account row:              {cfg['row_num']}  ({_posting_handle()})")
    print(f"  Account status:           {cfg.get('account_status') or '(not set)'}")
    print(f"  Mega upload folder:       {cfg['mega_upload_folder'] or '(not set!)'}")
    print(f"  Mega processed folder:    {cfg['mega_processed_folder'] or '(not set!)'}")
    print(f"  Post link (in-caption):   {cfg['link_display_text']} -> {cfg['link_url']}")
    print(f"  Post-type mix:            image {cfg['image_ratio']:.0%} / "
          f"video {cfg['video_ratio']:.0%} / previewLink {cfg['link_ratio']:.0%}")
    print(f"  Caption enabled:          {cfg['caption_enabled']}")
    print(f"  Hashtags on image posts:  {cfg['hashtags_enabled_image']}")
    print(f"  Hashtags on video posts:  {cfg['hashtags_enabled_video']}")
    print(f"  Hashtags on previewLink:  {cfg['hashtags_enabled_link']}")
    print(f"  Link on image posts:      {cfg['link_enabled_image']}")
    print(f"  Link on video posts:      {cfg['link_enabled_video']}")
    print(f"  Link inclusion rate:      {cfg['link_percentage']:.0%} of eligible image/video posts")
    print(f"  Max image size:           {cfg['max_image_bytes']/(1024*1024):.2f} MB")
    print(f"  previewLink auto-caption: {cfg['auto_caption_enabled_link']} (title+description when sheet has none)")
    print(f"  previewLink fetch timeout:{cfg['preview_timeout']}s")
    print(f"  previewLink max thumb:    {cfg['max_thumb_bytes']/(1024*1024):.2f} MB")
    print(f"  Loop interval:            {cfg['loop_interval_seconds']}s ({cfg['loop_interval_seconds']/60:.1f} min)")
    print(f"  Generate report:          {cfg['enable_report']}")
    if cfg["enable_report"]:
        print(f"  Report frequency:         {cfg['report_times_per_day']}x per 24h")
        print(f"  Top posts combined:       {cfg['top_posts_count']}")
        print(f"  Scan last N posts:        {cfg['top_posts_within']}")
    print(f"  Post-plan sheet (media):  {cfg['post_plan_sheet_name']}")
    print(f"  LinkPlan sheet:           {cfg['link_plan_sheet_name']}")
    print(f"  Config source:            Mega workbook ({STATE_REMOTE_FULL})")
    if cfg.get("has_lock_columns"):
        print(f"  Cross-repo lock:          enabled (owner={cfg.get('locked_by') or '—'}, "
              f"last heartbeat={cfg.get('locked_at') or '—'})")
    else:
        print("  Cross-repo lock:          disabled (add LOCKED_BY / LOCKED_AT columns to enable)")
    print(f"  Workbook lock TTL:        {WORKBOOK_LOCK_TTL_SECONDS}s "
          f"(acquire timeout {WORKBOOK_LOCK_ACQUIRE_TIMEOUT_SECONDS}s)")
    print("─────────────────────────────────────────────────")


# ═══════════════════════════════════════════════════════════════════════════
#  REPORT TAB
# ═══════════════════════════════════════════════════════════════════════════

def _now_str():
    return time.strftime("%Y-%m-%d %H:%M", time.gmtime()) + " UTC"

def _parse_report_ts(s):
    try:
        return time.mktime(time.strptime(s.replace(" UTC", ""), "%Y-%m-%d %H:%M"))
    except Exception:
        return None


def _last_report_for_handle(handle):
    wb = STORE.snapshot()
    if REPORT_TAB not in wb.sheetnames:
        return None, None
    ws = wb[REPORT_TAB]
    for r in range(ws.max_row, 1, -1):
        h = ws_cell(ws, r, 2)
        if h == handle:
            ts = ws_cell(ws, r, 1)
            followers = None
            f_raw = ws_cell(ws, r, 3)
            if f_raw:
                try:
                    followers = int(f_raw)
                except ValueError:
                    followers = None
            return ts, followers
    return None, None


def _report_due(handle, times_per_day):
    times_per_day = max(1, times_per_day)
    last_ts, _ = _last_report_for_handle(handle)
    if last_ts is None:
        return True
    last_epoch = _parse_report_ts(last_ts)
    if last_epoch is None:
        return True
    interval_seconds = 86400.0 / times_per_day
    return (time.time() - last_epoch) >= interval_seconds


def _append_report(rows):
    def _mut(wb):
        if REPORT_TAB not in wb.sheetnames:
            ws = wb.create_sheet(REPORT_TAB)
            ws.append(REPORT_HEADER)
        else:
            ws = wb[REPORT_TAB]
        for row in rows:
            ws.append(row)
        return True, None
    STORE.transact(_mut)


def _top_post_summary(client, handle, top_n, within):
    try:
        response = client.get_author_feed(actor=handle, limit=within)
    except Exception as exc:
        return f"(couldn't fetch posts: {exc})", 0

    posts = []
    for item in response.feed:
        if getattr(item, "reason", None) is not None:
            continue
        p       = item.post
        likes   = getattr(p, "like_count",   0) or 0
        reposts = getattr(p, "repost_count", 0) or 0
        replies = getattr(p, "reply_count",  0) or 0
        quotes  = getattr(p, "quote_count",  0) or 0
        try:
            text = p.record.text or ""
        except AttributeError:
            text = ""
        posts.append({
            "text": text, "likes": likes, "reposts": reposts,
            "replies": replies, "quotes": quotes,
            "engagement": likes + reposts + replies + quotes,
        })

    if not posts:
        return "(no posts found)", 0

    ranked = sorted(posts, key=lambda p: p["engagement"], reverse=True)[:max(1, top_n)]

    if len(ranked) == 1:
        p = ranked[0]
        preview = p["text"][:100] + ("…" if len(p["text"]) > 100 else "")
        return preview, p["engagement"]

    parts = []
    for i, p in enumerate(ranked, start=1):
        preview = p["text"][:60] + ("…" if len(p["text"]) > 60 else "")
        parts.append(f"{i}) {preview} ({p['engagement']})")
    return " | ".join(parts), ranked[0]["engagement"]


def generate_report(client, handle, cfg):
    """All Bluesky API calls happen with NO lock held; only the final
    `_append_report` touches the workbook lock, and only briefly."""
    if not _report_due(handle, cfg["report_times_per_day"]):
        print(f"Report for {handle} not due yet (limit: {cfg['report_times_per_day']}x/24h).")
        return
    try:
        profile = client.get_profile(actor=handle)
        total   = profile.followers_count or 0

        _, prev_followers = _last_report_for_handle(handle)
        prev   = prev_followers if prev_followers is not None else total
        gained = total - prev

        top_preview, top_engagement = _top_post_summary(
            client, handle, cfg["top_posts_count"], cfg["top_posts_within"]
        )

        row = [_now_str(), handle, total, gained, top_preview, top_engagement, "OK"]
        _append_report([row])
        print(f"Report logged for {handle}: {total} followers ({gained:+d} since last), "
              f"top post engagement={top_engagement}.")
    except Exception as exc:
        print(f"Warning: report generation failed: {exc}")


def run_report(client, handle, cfg):
    try:
        generate_report(client, handle, cfg)
    except Exception as exc:
        print(f"Warning: report generation failed: {exc}")


# ═══════════════════════════════════════════════════════════════════════════
#  ERROR TYPES
# ═══════════════════════════════════════════════════════════════════════════

class AccountTakenDownError(Exception):
    """Fatal — log to sheet, disable workflow."""

class NoMediaFoundError(Exception):
    """Clean exit (code 0) — nothing postable this cycle; keep schedule running."""

class NoPreviewError(Exception):
    """Link-preview metadata could not be fetched via cardyb or manual scrape."""


def log_account_problem(handle, status):
    try:
        _append_report([[_now_str(), handle, "", "", "", "", status]])
        print(f"Logged '{status}' for {handle}.")
    except Exception as exc:
        print(f"Warning: could not log account status: {exc}")


# ═══════════════════════════════════════════════════════════════════════════
#  ACCOUNT DISPLAY
# ═══════════════════════════════════════════════════════════════════════════

def print_target_account(handle):
    display = handle if handle.startswith("@") else f"@{handle}"
    print(f"Target Bluesky account: {display}")
    print(f"  (app password: {'loaded' if _cfg().get('app_pw') else 'MISSING!'})")


# ═══════════════════════════════════════════════════════════════════════════
#  HASHTAGS
# ═══════════════════════════════════════════════════════════════════════════

def get_account_hashtags():
    raw = _cfg().get("hashtags_raw", "")
    if raw:
        tags = [w.lstrip("#") for w in raw.split() if w.startswith("#")]
        if tags:
            return tags
    try:
        with open("hashtags.txt", "r", encoding="utf-8") as f:
            sets = [l.strip() for l in f if l.strip()]
        return [w.lstrip("#") for w in random.choice(sets).split() if w.startswith("#")] if sets else []
    except FileNotFoundError:
        return []


# ═══════════════════════════════════════════════════════════════════════════
#  LINK-IN-POST DECISION
# ═══════════════════════════════════════════════════════════════════════════

def should_add_link(kind):
    cfg     = _cfg()
    enabled = cfg["link_enabled_image"] if kind == "image" else cfg["link_enabled_video"]
    if not enabled:
        return False
    return random.random() < cfg["link_percentage"]


# ═══════════════════════════════════════════════════════════════════════════
#  POST-TYPE SELECTION
# ═══════════════════════════════════════════════════════════════════════════

def choose_media_kind():
    cfg = _cfg()
    return random.choices(
        ["image", "video", "previewLink"],
        weights=[cfg["image_ratio"], cfg["video_ratio"], cfg["link_ratio"]],
        k=1,
    )[0]


# ═══════════════════════════════════════════════════════════════════════════
#  POST-PLAN SHEET (File Name + Caption + Status) — reads are unlocked,
#  the only write (mark_posted) is a short transaction.
# ═══════════════════════════════════════════════════════════════════════════

_post_plan_cache          = None
_post_plan_status_col_idx = None   # 1-based col index


def get_post_plan_tab_name():
    return _cfg()["post_plan_sheet_name"]


def load_post_plan(force_refresh=False):
    global _post_plan_cache, _post_plan_status_col_idx
    if _post_plan_cache is not None and not force_refresh:
        return _post_plan_cache

    tab_name = get_post_plan_tab_name()
    wb = STORE.snapshot()
    if tab_name not in wb.sheetnames:
        print(f"Warning: post-plan sheet '{tab_name}' not found in the workbook — "
              f"add a tab with that name (File Name / Caption / Status columns).")
        _post_plan_cache = {}
        return _post_plan_cache

    ws = wb[tab_name]
    header = ws_header_map(ws)

    def ci(*names):
        for n in names:
            if n.upper() in header:
                return header[n.upper()]
        return None

    file_idx    = ci("file name", "filename", "file")
    caption_idx = ci("caption", "captions")
    status_idx  = ci("status")
    _post_plan_status_col_idx = status_idx

    if file_idx is None or caption_idx is None:
        print(f"Warning: post-plan needs 'File Name' and 'Caption' columns. Found: {list(header)}")
        _post_plan_cache = {}
        return _post_plan_cache
    if status_idx is None:
        print("Warning: no 'Status' column — posted files won't be tracked.")

    plan_exact = {}
    plan_lower = {}
    already    = 0
    for excel_row in range(2, ws.max_row + 1):
        fname   = ws_cell(ws, excel_row, file_idx)
        caption = ws_cell(ws, excel_row, caption_idx)
        status  = ws_cell(ws, excel_row, status_idx) if status_idx else ""
        if not fname:
            continue
        entry = {"caption": caption, "row": excel_row, "status": status}
        plan_exact[fname]         = entry
        plan_lower[fname.lower()] = entry
        if status.lower() == POSTED_STATUS_VALUE:
            already += 1

    print(f"Loaded {len(plan_exact)} post-plan rows ({already} already posted).")
    _post_plan_cache = {"exact": plan_exact, "lower": plan_lower}
    return _post_plan_cache


def find_plan_entry(plan, filename):
    exact = plan.get("exact", {})
    lower = plan.get("lower", {})
    return (
        exact.get(filename)
        or lower.get(filename.lower())
        or lower.get(os.path.splitext(filename.lower())[0])
    )


def mark_posted(filename, row_number, retries=3):
    global _post_plan_cache
    if _post_plan_status_col_idx is None:
        print(f"Warning: no 'Status' column — cannot mark '{filename}' as posted.")
        return

    tab_name = get_post_plan_tab_name()

    def _mut(wb):
        ws = wb[tab_name]
        ws_set_cell(ws, row_number, _post_plan_status_col_idx, POSTED_STATUS_VALUE)
        return True, None

    for attempt in range(1, retries + 1):
        try:
            STORE.transact(_mut)
            if _post_plan_cache:
                for d in (_post_plan_cache.get("exact", {}), _post_plan_cache.get("lower", {})):
                    for entry in d.values():
                        if entry["row"] == row_number:
                            entry["status"] = POSTED_STATUS_VALUE
            print(f"Marked '{filename}' row {row_number} as posted.")
            return
        except Exception as exc:
            if attempt < retries:
                wait = 2 ** attempt
                print(f"  mark_posted attempt {attempt}/{retries} failed ({exc}); retrying in {wait}s…")
                time.sleep(wait)
            else:
                print(f"ERROR: could not mark '{filename}' as posted after {retries} attempts: {exc}")
                print("  Post was successful — file will be moved. Row may need manual update.")


# ═══════════════════════════════════════════════════════════════════════════
#  LINKPLAN SHEET (URL + Caption + Status)
# ═══════════════════════════════════════════════════════════════════════════

def get_link_plan_tab_name():
    return _cfg()["link_plan_sheet_name"]


def load_link_plan():
    tab_name = get_link_plan_tab_name()
    wb = STORE.snapshot()
    if tab_name not in wb.sheetnames:
        return []

    ws = wb[tab_name]
    header = ws_header_map(ws)

    def ci(*names):
        for n in names:
            if n.upper() in header:
                return header[n.upper()]
        return None

    url_idx    = ci("url")
    cap_idx    = ci("caption")
    status_idx = ci("status")
    if url_idx is None:
        raise RuntimeError(f"'{tab_name}' needs a 'URL' column.")

    out = []
    for excel_row in range(2, ws.max_row + 1):
        url = ws_cell(ws, excel_row, url_idx)
        if not url:
            continue
        caption = ws_cell(ws, excel_row, cap_idx) if cap_idx else ""
        status  = ws_cell(ws, excel_row, status_idx) if status_idx else ""
        out.append({
            "url": url, "caption": caption, "status": status,
            "row": excel_row, "status_col": status_idx,
        })
    return out


def pick_next_url():
    plan = load_link_plan()
    for entry in plan:
        s = entry["status"].lower()
        if s == POSTED_STATUS_VALUE or s.startswith(CLAIM_PREFIX.lower()):
            continue
        return entry
    return None


def claim_url_row(entry):
    if entry["status_col"] is None:
        return True
    tab_name  = get_link_plan_tab_name()
    claim_val = f"{CLAIM_PREFIX}{RUN_TAG}"

    def _mut(wb):
        ws = wb[tab_name]
        current = ws_cell(ws, entry["row"], entry["status_col"])
        if current.lower() == POSTED_STATUS_VALUE or current.lower().startswith(CLAIM_PREFIX.lower()):
            return False, False   # someone else got there first
        ws_set_cell(ws, entry["row"], entry["status_col"], claim_val)
        return True, True

    return STORE.transact(_mut)


def mark_url_posted(entry):
    if entry["status_col"] is None:
        return
    tab_name = get_link_plan_tab_name()

    def _mut(wb):
        ws = wb[tab_name]
        ws_set_cell(ws, entry["row"], entry["status_col"], POSTED_STATUS_VALUE)
        return True, None

    STORE.transact(_mut)


def release_url_claim(entry):
    if entry["status_col"] is None:
        return
    tab_name = get_link_plan_tab_name()

    def _mut(wb):
        ws = wb[tab_name]
        ws_set_cell(ws, entry["row"], entry["status_col"], "")
        return True, None

    try:
        STORE.transact(_mut)
    except Exception as exc:
        print(f"Warning: could not release claim on LinkPlan row {entry['row']}: {exc}")


# ═══════════════════════════════════════════════════════════════════════════
#  LINK PREVIEW
# ═══════════════════════════════════════════════════════════════════════════

def fetch_link_metadata(url, timeout=20):
    last_exc = None
    for attempt in range(1, LINK_PREVIEW_MAX_RETRIES + 1):
        try:
            resp = requests.get(
                CARDYB_EXTRACT_URL,
                params={"url": url},
                headers=REQUEST_HEADERS,
                timeout=timeout,
            )
            resp.raise_for_status()
            data = resp.json()
            return {
                "title": (data.get("title") or url)[:300],
                "description": (data.get("description") or "")[:1000],
                "image": data.get("image") or None,
                "final_url": url,
            }
        except (requests.exceptions.RequestException, ValueError) as exc:
            last_exc = exc
            print(f"Attempt {attempt}/{LINK_PREVIEW_MAX_RETRIES} to fetch via cardyb failed: {exc}")
            if attempt < LINK_PREVIEW_MAX_RETRIES:
                delay = LINK_PREVIEW_RETRY_DELAY * (2 ** (attempt - 1))
                print(f"Retrying in {delay}s…")
                time.sleep(delay)

    print(f"cardyb extraction failed after {LINK_PREVIEW_MAX_RETRIES} attempts ({last_exc}); "
          f"falling back to manual scrape.")
    return _fetch_link_metadata_manual(url, timeout)


def _fetch_link_metadata_manual(url, timeout=20):
    last_exc = None
    for attempt in range(1, LINK_PREVIEW_MAX_RETRIES + 1):
        try:
            resp = requests.get(url, headers=REQUEST_HEADERS, timeout=timeout, allow_redirects=True)
            resp.raise_for_status()
            return _parse_link_metadata(resp)
        except requests.exceptions.RequestException as exc:
            last_exc = exc
            print(f"Attempt {attempt}/{LINK_PREVIEW_MAX_RETRIES} to manually fetch {url} failed: {exc}")
            if attempt < LINK_PREVIEW_MAX_RETRIES:
                delay = LINK_PREVIEW_RETRY_DELAY * (2 ** (attempt - 1))
                print(f"Retrying in {delay}s…")
                time.sleep(delay)

    raise NoPreviewError(
        f"Failed to fetch {url} after {LINK_PREVIEW_MAX_RETRIES} attempts (cardyb + manual)"
    ) from last_exc


def _parse_link_metadata(resp):
    soup = BeautifulSoup(resp.text, "html.parser")
    final_url = resp.url

    def meta(*props):
        for prop in props:
            tag = soup.find("meta", property=prop) or soup.find("meta", attrs={"name": prop})
            if tag and tag.get("content"):
                return tag["content"].strip()
        return None

    title = meta("og:title", "twitter:title") or (
        soup.title.string.strip() if soup.title and soup.title.string else resp.url
    )
    description = meta("og:description", "twitter:description", "description") or ""
    raw_image = meta("og:image", "og:image:url", "twitter:image")
    image = urljoin(final_url, raw_image) if raw_image else None

    return {"title": title[:300], "description": description[:1000], "image": image, "final_url": final_url}


def upload_link_thumbnail(client, image_url, referer, max_bytes, timeout=20):
    if not image_url:
        print("No preview image found — posting without a thumbnail.")
        return None

    headers = {**REQUEST_HEADERS, "Referer": referer}
    last_exc = None
    for attempt in range(1, LINK_PREVIEW_MAX_RETRIES + 1):
        try:
            img_resp = requests.get(image_url, headers=headers, timeout=timeout)
            img_resp.raise_for_status()
            content_type = img_resp.headers.get("Content-Type", "")
            if "image" not in content_type:
                print(f"Warning: fetched image URL did not return an image (Content-Type: {content_type!r})")
                return None

            data = img_resp.content
            if len(data) > max_bytes:
                data = _compress_link_thumb(data, max_bytes)
                if data is None:
                    return None

            upload = client.upload_blob(data)
            return upload.blob
        except Exception as exc:  # noqa: BLE001
            last_exc = exc
            print(f"Attempt {attempt}/{LINK_PREVIEW_MAX_RETRIES} to fetch/upload thumbnail failed: {exc}")
            if attempt < LINK_PREVIEW_MAX_RETRIES:
                delay = LINK_PREVIEW_RETRY_DELAY * (2 ** (attempt - 1))
                print(f"Retrying in {delay}s…")
                time.sleep(delay)

    print(f"Warning: thumbnail could not be fetched/uploaded after all retries ({last_exc}); posting without one.")
    return None


def _compress_link_thumb(data, max_bytes):
    try:
        from PIL import Image
        img = Image.open(io.BytesIO(data))
        if img.mode in ("RGBA", "P", "LA"):
            img = img.convert("RGB")
        for q in range(85, 20, -10):
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=q, optimize=True)
            if buf.tell() <= max_bytes:
                return buf.getvalue()
        return buf.getvalue()
    except Exception as exc:
        print(f"Warning: could not compress thumbnail: {exc}")
        return None


MAX_POST_GRAPHEMES = 300

def build_link_caption_text(caption, tags, fallback_url=None):
    text = _URL_RE.sub("", caption or "").strip(" \t\r")
    if fallback_url:
        text = f"{text}\n{fallback_url}".strip(" \t\r") if text else fallback_url

    tb = TextBuilder()
    if text:
        tb.text(text)
    if tags:
        if text:
            tb.text("\n\n")
        for i, tag in enumerate(tags):
            tb.tag(f"#{tag}", tag)
            if i < len(tags) - 1:
                tb.text(" ")

    plain = tb.build_text()
    if len(plain) > MAX_POST_GRAPHEMES:
        hashtag_block = ("\n\n" + " ".join(f"#{t}" for t in tags)) if tags else ""
        budget = MAX_POST_GRAPHEMES - len(hashtag_block)
        trimmed = (text[:max(0, budget - 1)].rstrip() + "…") if budget > 0 else ""
        tb = TextBuilder()
        if trimmed:
            tb.text(trimmed)
        if tags:
            if trimmed:
                tb.text("\n\n")
            for i, tag in enumerate(tags):
                tb.tag(f"#{tag}", tag)
                if i < len(tags) - 1:
                    tb.text(" ")
    return tb


def build_external_embed(client, preview, max_thumb_bytes, timeout):
    thumb_blob = upload_link_thumbnail(
        client, preview["image"], referer=preview["final_url"],
        max_bytes=max_thumb_bytes, timeout=timeout,
    )
    return models.AppBskyEmbedExternal.Main(
        external=models.AppBskyEmbedExternal.External(
            uri=preview["final_url"],
            title=preview["title"],
            description=preview["description"],
            thumb=thumb_blob,
        )
    )


def compose_fallback_caption(preview):
    if not preview:
        return ""
    title = (preview.get("title") or "").strip()
    description = (preview.get("description") or "").strip()
    parts = [p for p in (title, description) if p]
    if not parts:
        return ""
    return "\n" + "\n".join(parts)


def post_link_card(client, url, caption, tags, timeout, max_thumb_bytes, auto_caption_enabled=True):
    print(f"[previewLink] Fetching preview for: {url}")
    preview = None
    embed = None
    try:
        preview = fetch_link_metadata(url, timeout)
        print(f"  title: {preview['title']!r}")
        embed = build_external_embed(client, preview, max_thumb_bytes, timeout)
    except Exception as exc:
        print(f"Warning: preview fetch failed ({exc}); posting as plain link instead.")

    used_auto_caption = False
    effective_caption = caption
    if not effective_caption and preview and auto_caption_enabled:
        effective_caption = compose_fallback_caption(preview)
        used_auto_caption = bool(effective_caption)
        if used_auto_caption:
            print("No Caption in sheet — using title + description from the preview instead.")
    elif not effective_caption and preview and not auto_caption_enabled:
        print("No Caption in sheet and previewLink auto-caption is off — posting without a caption.")

    tb = build_link_caption_text(effective_caption, tags, fallback_url=(url if preview is None else None))
    client.send_post(text=tb, embed=embed)

    posted_url = preview["final_url"] if preview else url
    caption_source = "auto (title+description)" if used_auto_caption else ("sheet" if caption else "no")
    print(f"✓ Posted {'link card' if embed else 'plain link'} for {posted_url} "
          f"(caption={caption_source}, tags={len(tags)})")


# ═══════════════════════════════════════════════════════════════════════════
#  MEGA.NZ HELPERS (via rclone) — image/video media, unchanged
# ═══════════════════════════════════════════════════════════════════════════

_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tiff", ".avif", ".heic"}
_VIDEO_EXTS = {".mp4", ".mov", ".avi", ".mkv", ".webm", ".m4v", ".flv", ".wmv", ".3gp", ".ts"}


def _kind_from_filename(filename):
    ext = os.path.splitext(filename.lower())[1]
    if ext in _IMAGE_EXTS:
        return "image"
    if ext in _VIDEO_EXTS:
        return "video"
    return None


def fetch_media_matching_plan(preferred_kind, plan):
    cfg           = _cfg()
    upload_folder = cfg["mega_upload_folder"]
    if not upload_folder:
        raise RuntimeError("MEGA_UPLOAD_FOLDER is empty in credentials sheet.")

    remote_folder = f"{RCLONE_REMOTE_NAME}:{upload_folder}"
    files = rclone_list_files(remote_folder)
    candidates = [f for f in files if not f.startswith(CLAIM_PREFIX)
                  and _kind_from_filename(f) == preferred_kind]

    counters = {"claim": 0, "plan": 0, "posted": 0}

    for name in candidates:
        entry = find_plan_entry(plan, name)
        if entry is None:
            counters["plan"] += 1
            continue
        if entry["status"].lower() == POSTED_STATUS_VALUE:
            counters["posted"] += 1
            continue

        print(f"Found {preferred_kind}: '{name}'")
        claimed_name = rclone_claim(remote_folder, name)
        if claimed_name is None:
            counters["claim"] += 1
            continue
        print(f"Claimed as '{claimed_name}'.")

        local_path = f"/tmp/{name}"
        if not rclone_download(remote_folder, claimed_name, local_path):
            print(f"Warning: download failed for claimed file '{claimed_name}' — releasing claim.")
            rclone_move(f"{remote_folder}/{claimed_name}", f"{remote_folder}/{name}")
            continue

        file_info = {"original_name": name, "claimed_name": claimed_name}
        return file_info, local_path, preferred_kind, entry["caption"], entry["row"]

    print(f"No {preferred_kind} match: "
          f"{counters['plan']} not in plan, {counters['posted']} already posted, "
          f"{counters['claim']} claimed by another run.")
    return None, None, None, None, None


def compress_image_under_limit(local_path):
    from PIL import Image
    max_bytes = _cfg()["max_image_bytes"]
    orig = os.path.getsize(local_path)
    if orig <= max_bytes:
        print(f"Image {orig/1024:.0f} KB — no compression needed.")
        return local_path
    img = Image.open(local_path)
    if img.mode in ("RGBA", "P", "LA"):
        img = img.convert("RGB")
    for q in range(90, 20, -10):
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=q, optimize=True)
        if buf.tell() <= max_bytes:
            with open(local_path, "wb") as f:
                f.write(buf.getvalue())
            print(f"Compressed {orig/1024:.0f} KB → {buf.tell()/1024:.0f} KB (q={q}).")
            return local_path
    w, h = img.size
    scale = 0.9
    while scale > 0.3:
        r = img.resize((max(1, int(w * scale)), max(1, int(h * scale))), Image.LANCZOS)
        buf = io.BytesIO()
        r.save(buf, format="JPEG", quality=70, optimize=True)
        if buf.tell() <= max_bytes:
            with open(local_path, "wb") as f:
                f.write(buf.getvalue())
            print(f"Resized+compressed → {buf.tell()/1024:.0f} KB.")
            return local_path
        scale -= 0.1
    with open(local_path, "wb") as f:
        f.write(buf.getvalue())
    print(f"Warning: best-effort compression = {buf.tell()/1024:.0f} KB.")
    return local_path


def move_file_to_processed(claimed_name, original_name):
    cfg = _cfg()
    remote_upload    = f"{RCLONE_REMOTE_NAME}:{cfg['mega_upload_folder']}"
    remote_processed = f"{RCLONE_REMOTE_NAME}:{cfg['mega_processed_folder']}"
    ok = rclone_move(f"{remote_upload}/{claimed_name}", f"{remote_processed}/{original_name}")
    print("Moved to processed folder on Mega." if ok else
          "Warning: failed to move file to processed folder on Mega — check manually.")
    return ok


def release_claim(claimed_name, original_name):
    cfg = _cfg()
    remote_upload = f"{RCLONE_REMOTE_NAME}:{cfg['mega_upload_folder']}"
    ok = rclone_move(f"{remote_upload}/{claimed_name}", f"{remote_upload}/{original_name}")
    print(f"Released claim on '{original_name}'." if ok else
          f"Warning: could not release claim for '{original_name}'.")
    return ok


# ═══════════════════════════════════════════════════════════════════════════
#  IMAGE/VIDEO POST BUILDING
# ═══════════════════════════════════════════════════════════════════════════

def build_post_from_caption(caption, tags, add_link):
    cfg  = _cfg()
    text = replace_mentions(caption) if caption else ""

    def _assemble(caption_text):
        tb = TextBuilder()
        if add_link:
            m = _URL_RE.search(caption_text)
            if m:
                before = caption_text[:m.start()].rstrip()
                after  = _URL_RE.sub("", caption_text[m.end():]).strip()
                if before:
                    tb.text(before + " ")
                tb.link(cfg["link_display_text"], cfg["link_url"])
                if after:
                    tb.text(" " + after)
            else:
                if caption_text:
                    tb.text(caption_text)
                    tb.text("\n\n")
                tb.link(cfg["link_display_text"], cfg["link_url"])
        else:
            text_no_url = _URL_RE.sub("", caption_text).strip()
            if text_no_url:
                tb.text(text_no_url)

        if tags:
            tb.text("\n\n")
            for i, tag in enumerate(tags):
                tb.tag(f"#{tag}", tag)
                if i < len(tags) - 1:
                    tb.text(" ")
        return tb

    tb    = _assemble(text)
    plain = tb.build_text()

    if len(plain) > MAX_POST_GRAPHEMES:
        lo, hi, best_text = 0, len(text), ""
        while lo <= hi:
            mid   = (lo + hi) // 2
            trial = text[:mid].rstrip()
            if mid < len(text):
                trial += "…"
            if len(_assemble(trial).build_text()) <= MAX_POST_GRAPHEMES:
                best_text = trial
                lo = mid + 1
            else:
                hi = mid - 1
        print(f"Caption too long for post limit ({len(plain)} > {MAX_POST_GRAPHEMES}); "
              f"trimmed caption to fit.")
        tb = _assemble(best_text)

    return tb


def post_to_bluesky(client, media_name, local_path, kind, caption, tags, add_link):
    tb = build_post_from_caption(caption, tags, add_link)
    if kind == "video":
        with open(local_path, "rb") as f:
            client.send_video(text=tb, video=f.read(), video_alt=media_name)
    else:
        with open(local_path, "rb") as f:
            client.send_image(text=tb, image=f.read(), image_alt=media_name)

    preview = replace_mentions(caption or "")
    if add_link:
        m = _URL_RE.search(preview)
        if m:
            preview = (preview[:m.start()].rstrip()
                       + f" [{_cfg()['link_display_text']}]"
                       + _URL_RE.sub("", preview[m.end():]).strip())
        else:
            preview = (preview + f" [{_cfg()['link_display_text']}]").strip()
    else:
        preview = _URL_RE.sub("", preview).strip()
    print(f"✓ Posted {kind}: {preview!r} (link={'yes' if add_link else 'no'})")
    if tags:
        print(f"  Tags: {' '.join('#'+t for t in tags)}")


# ═══════════════════════════════════════════════════════════════════════════
#  DISCOVER MODE
# ═══════════════════════════════════════════════════════════════════════════

def run_discover():
    bootstrap_state()

    wb = STORE.snapshot()
    if CREDS_TAB not in wb.sheetnames:
        print(f"::error::'{CREDS_TAB}' tab not found in the workbook.")
        sys.exit(1)
    ws = wb[CREDS_TAB]
    header = ws_header_map(ws)

    def hidx(*names):
        for n in names:
            if n.upper() in header:
                return header[n.upper()]
        return None

    handle_col = hidx("BSKY_HANDLE")
    status_col = hidx("ACCOUNT_STATUS")

    if handle_col is None:
        print(f"::error::'{CREDS_TAB}' needs a 'BSKY_HANDLE' column.")
        sys.exit(1)

    force_row = get_env("FORCE_ACCOUNT_ROW", required=False)
    if force_row:
        try:
            eligible = [max(1, int(force_row))]
        except ValueError:
            print(f"::error::FORCE_ACCOUNT_ROW={force_row!r} is not a valid row number.")
            sys.exit(1)
        print(f"FORCE_ACCOUNT_ROW set — running only row {eligible[0]} "
              f"(eligibility filter and MAX_ACCOUNTS_PER_RUN cap skipped).")
    else:
        eligible = []
        for excel_row in range(2, ws.max_row + 1):
            handle = ws_cell(ws, excel_row, handle_col)
            if not handle:
                continue
            status = ws_cell(ws, excel_row, status_col).lower() if status_col else ""
            if any(marker in status for marker in SKIP_STATUS_MARKERS):
                print(f"Skipping row {excel_row - 1} ({handle}) — status: {status!r}")
                continue
            eligible.append(excel_row - 1)

        if not eligible:
            print(f"::error::No eligible account rows in '{CREDS_TAB}' "
                  f"(all rows are empty or flagged banned/suspended).")
            sys.exit(1)

        limit_raw = get_env("MAX_ACCOUNTS_PER_RUN", required=False)
        if not limit_raw:
            settings = load_global_settings()
            limit_raw = settings.get("MAX_ACCOUNTS_PER_RUN", "")

        if limit_raw:
            try:
                limit = max(1, int(limit_raw))
                if limit < len(eligible):
                    print(f"Capping this run to the first {limit} of "
                          f"{len(eligible)} eligible accounts (MAX_ACCOUNTS_PER_RUN={limit}).")
                eligible = eligible[:limit]
            except ValueError:
                print(f"Warning: MAX_ACCOUNTS_PER_RUN={limit_raw!r} is not a valid "
                      f"number — running all eligible rows.")
        else:
            print("MAX_ACCOUNTS_PER_RUN not set — running all eligible accounts.")

    print(f"Account rows for this workflow run: {eligible}")
    rows_csv = ",".join(str(r) for r in eligible)

    gh_output = os.environ.get("GITHUB_OUTPUT")
    if not gh_output:
        raise RuntimeError("GITHUB_OUTPUT is not set — must be run inside a GitHub Actions step.")
    with open(gh_output, "a") as f:
        f.write(f"rows={rows_csv}\n")
        f.write(f"count={len(eligible)}\n")


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN CYCLE
# ═══════════════════════════════════════════════════════════════════════════

def run_once():
    cfg = refresh_account_config()

    if not try_acquire_account_lock():
        raise AccountLockedElsewhereError(
            f"Account row {ACCOUNT_ROW} is locked by another repo right now."
        )

    handle = cfg["handle"]

    print_target_account(handle)
    client = Client()
    try:
        client.login(handle, cfg["app_pw"])
    except Exception as exc:
        err = str(exc)
        if "AccountTakedown" in err or "AccountSuspended" in err:
            raise AccountTakenDownError(f"Account {handle} taken down/suspended.") from exc
        if "AuthenticationRequired" in err or "Invalid identifier or password" in err:
            raise AccountTakenDownError(
                f"Auth failed for {handle} — check BSKY_HANDLE / BSKY_APP_PW in Credentials row {ACCOUNT_ROW}."
            ) from exc
        raise

    _write_account_status("Active")

    if cfg["enable_report"]:
        run_report(client, handle, cfg)

    preferred = choose_media_kind()
    print(f"Post-type chosen for this cycle: {preferred}")

    # ── previewLink path ────────────────────────────────────────────────
    if preferred == "previewLink":
        entry = pick_next_url()
        if entry is None:
            print("No unposted rows left in LinkPlan — falling back to image/video this cycle.")
            preferred = random.choice(["image", "video"])
        else:
            if not claim_url_row(entry):
                print("Lost claim race on this LinkPlan row; will try again next cycle.")
                return

            tags = get_account_hashtags() if cfg["hashtags_enabled_link"] else []
            try:
                post_link_card(
                    client, entry["url"], entry["caption"], tags,
                    cfg["preview_timeout"], cfg["max_thumb_bytes"],
                    auto_caption_enabled=cfg["auto_caption_enabled_link"],
                )
            except Exception as exc:
                err = str(exc)
                release_url_claim(entry)
                if "AccountTakedown" in err or "AccountSuspended" in err:
                    raise AccountTakenDownError(f"Account {handle} taken down mid-cycle.") from exc
                print(f"previewLink post failed for {entry['url']} — claim released: {exc}")
                raise

            mark_url_posted(entry)
            return

    # ── image/video path ────────────────────────────────────────────────
    plan = load_post_plan()
    if not plan:
        raise NoMediaFoundError("Post-plan sheet has no usable rows.")

    fallback = "video" if preferred == "image" else "image"

    file, path, kind, caption, row_num = fetch_media_matching_plan(preferred, plan)
    if not file:
        print(f"No {preferred} matched; trying {fallback}.")
        file, path, kind, caption, row_num = fetch_media_matching_plan(fallback, plan)

    if not file:
        raise NoMediaFoundError("No unposted Mega file matching the post-plan sheet.")

    original_name = file["original_name"]
    claimed_name  = file["claimed_name"]

    try:
        if kind == "image":
            path = compress_image_under_limit(path)

        cfg = _cfg()
        hashtags_on = cfg["hashtags_enabled_image"] if kind == "image" else cfg["hashtags_enabled_video"]
        tags = get_account_hashtags() if hashtags_on else []
        add_link = should_add_link(kind)
        caption_to_use = caption if cfg.get("caption_enabled", True) else ""

        post_to_bluesky(client, original_name, path, kind, caption_to_use, tags, add_link)

    except Exception as exc:
        err = str(exc)
        if "AccountTakedown" in err or "AccountSuspended" in err:
            release_claim(claimed_name, original_name)
            raise AccountTakenDownError(f"Account {handle} taken down mid-cycle.") from exc
        release_claim(claimed_name, original_name)
        print(f"Post failed — claim released, file stays in upload folder.")
        raise

    mark_posted(original_name, row_num)
    move_file_to_processed(claimed_name, original_name)
    try:
        os.remove(path)
    except OSError:
        pass


def main():
    global ACCOUNT_ROW
    try:
        bootstrap_state()
        ensure_settings_defaults()
        ACCOUNT_ROW = resolve_account_row()
        load_account_config()
    except Exception as exc:
        print(f"\n{'='*60}\nFATAL: {exc}\n{'='*60}\n")
        sys.exit(1)

    print_config_summary()
    print(f"Starting loop. Loop interval and post-type mix are read from the "
          f"Settings tab of the Mega workbook and re-checked at the start of "
          f"every cycle — edit them there any time, no redeploy needed.")

    while True:
        cycle_start = time.time()
        try:
            run_once()
        except AccountLockedElsewhereError as exc:
            print(f"\n{'='*60}\n{exc}\nSkipping — schedule keeps running.\n{'='*60}\n")
            sys.exit(0)
        except NoMediaFoundError as exc:
            print(f"\n{'='*60}\nNO MEDIA: {exc}\nStopping — schedule keeps running.\n{'='*60}\n")
            sys.exit(0)
        except AccountTakenDownError as exc:
            handle  = (_account_config or {}).get("handle", "unknown")
            err_str = str(exc)
            reason  = ("🔑 AUTH FAILED — check handle/app-password in sheet"
                       if "Auth failed" in err_str or "app password" in err_str
                       else "⛔ ACCOUNT TAKEN DOWN / BANNED")
            print(f"\n{'='*60}\n{err_str}\n→ {reason}\n{'='*60}\n")
            _write_account_status(reason)
            log_account_problem(handle, status=reason)
            with open("ACCOUNT_BANNED", "w") as f:
                f.write(f"{handle}: {reason}\n")
            sys.exit(1)
        except Exception as exc:
            print(f"Error during cycle: {exc}")

        loop_interval = (_account_config or {}).get("loop_interval_seconds", DEFAULT_LOOP_INTERVAL_SECONDS)
        elapsed   = time.time() - cycle_start
        sleep_for = max(0, loop_interval - elapsed)
        print(f"Cycle done in {elapsed:.1f}s. Sleeping {sleep_for:.1f}s "
              f"(interval={loop_interval}s from Settings tab)…")
        time.sleep(sleep_for)


if __name__ == "__main__":
    if "--discover" in sys.argv:
        run_discover()
    else:
        main()
